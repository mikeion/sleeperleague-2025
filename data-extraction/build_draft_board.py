"""Build the merged draft board for draft.html.

Combines four sources:
  - DraftSheets_2026.xlsx   projections (LOW/AVG/HIGH), projected missed games
  - Boris Chen              half-PPR tiers (gaussian mixture over FantasyPros ECR)
  - Sleeper projections     half-PPR ADP
  - Sleeper player DB       player_ids, so live draft picks can be matched

VBD is recomputed for THIS league (14 teams, 2 flex) rather than using the
workbook's values, which are baselined for a 12-team league with one flex.
"""
import json, re, sys, unicodedata, urllib.request
from pathlib import Path

import openpyxl

XLSX = Path.home() / "Downloads" / "DraftSheets_2026 (Copy).xlsx"
OUT = Path(__file__).parent.parent / "html5up-landed" / "assets" / "data" / "draft_board.json"
CHEN = "https://s3-us-west-1.amazonaws.com/fftiers/out/text_{}.txt"
SLEEPER_PROJ = ("https://api.sleeper.app/projections/nfl/2026?season_type=regular"
                "&position[]=QB&position[]=RB&position[]=WR&position[]=TE"
                "&position[]=K&position[]=DEF&order_by=pts_half_ppr")
SLEEPER_PLAYERS = "https://api.sleeper.app/v1/players/nfl"

# League: 14 teams, QB/RB/RB/WR/WR/TE/FLEX/WRRB_FLEX/K/DEF + 4 bench
TEAMS = 14
BASE = {"QB": 1, "RB": 2, "WR": 2, "TE": 1, "K": 1, "DEF": 1}
FLEX = [("FLEX", {"RB", "WR", "TE"}), ("WRRB_FLEX", {"RB", "WR"})]


def norm(name):
    n = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode()
    n = re.sub(r"\b(Jr|Sr|II|III|IV|V)\b\.?", "", n)
    return re.sub(r"[^a-z]", "", n.lower())


def get(url, binary=False):
    with urllib.request.urlopen(url, timeout=120) as r:
        raw = r.read()
    return raw if binary else json.loads(raw)


def chen_tiers():
    tiers = {}
    for pos, key in [("RB", "RB-HALF"), ("WR", "WR-HALF"), ("TE", "TE"), ("QB", "QB")]:
        txt = get(CHEN.format(key), binary=True).decode()
        for line in txt.splitlines():
            m = re.match(r"Tier (\d+): (.*)", line.strip())
            if m:
                for nm in m.group(2).split(","):
                    tiers[(pos, norm(nm))] = int(m.group(1))
    return tiers


def workbook_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Aggregate"]
    rows = []
    for pos, c0 in {"QB": 2, "RB": 16, "WR": 30, "TE": 44}.items():
        for r in range(3, ws.max_row + 1):
            nm = ws.cell(r, c0).value
            if not nm:
                continue
            cell = lambda o: ws.cell(r, c0 + o).value
            avg = cell(5)
            if not isinstance(avg, (int, float)):
                continue
            rows.append({"pos": pos, "name": str(nm).strip(), "team": cell(1),
                         "bye": cell(2), "low": cell(4), "proj": avg, "high": cell(6),
                         "miss": cell(11)})
    return rows


def replacement_levels(by_pos):
    """Greedy fill of every starting slot in the league to find replacement rank."""
    counts = {p: BASE[p] * TEAMS for p in BASE}
    pools = {p: sorted(v, key=lambda x: -x["proj"]) for p, v in by_pos.items()}
    used = {p: counts.get(p, 0) for p in pools}
    for _slot, eligible in FLEX:
        for _ in range(TEAMS):
            best, bp = None, None
            for p in eligible:
                nxt = pools.get(p, [])
                if used.get(p, 0) < len(nxt):
                    cand = nxt[used[p]]
                    if best is None or cand["proj"] > best["proj"]:
                        best, bp = cand, p
            if bp:
                used[bp] += 1
    repl = {}
    for p, pool in pools.items():
        idx = min(used.get(p, 0), len(pool) - 1)
        repl[p] = pool[idx]["proj"] if pool else 0.0
    return repl, used


def main():
    if not XLSX.exists():
        sys.exit(f"missing workbook: {XLSX}")
    tiers = chen_tiers()
    proj = get(SLEEPER_PROJ)
    players = get(SLEEPER_PLAYERS)

    # sleeper lookup: (pos, normalized name) -> player_id
    pid_by = {}
    for pid, p in players.items():
        pos = p.get("position")
        nm = p.get("full_name") or p.get("last_name")
        if pos and nm:
            pid_by.setdefault((pos, norm(nm)), pid)
    adp, sproj = {}, {}
    for r in proj:
        s, p = r.get("stats") or {}, r.get("player") or {}
        pos = p.get("position")
        pid = r.get("player_id")
        if not pos or not pid:
            continue
        a = s.get("adp_half_ppr")
        if a and a < 999:
            adp[pid] = a
        if s.get("pts_half_ppr") is not None:
            sproj[pid] = s["pts_half_ppr"]

    board, unmatched = [], []
    for row in workbook_rows():
        pid = pid_by.get((row["pos"], norm(row["name"])))
        if not pid:
            unmatched.append(f"{row['name']} ({row['pos']})")
            continue
        board.append({**row, "id": pid, "tier": tiers.get((row["pos"], norm(row["name"]))),
                      "adp": adp.get(pid)})

    # K and DEF are absent from the workbook; take them from Sleeper projections
    have = {b["id"] for b in board}
    for r in proj:
        p, pid = r.get("player") or {}, r.get("player_id")
        pos = p.get("position")
        if pos not in ("K", "DEF") or pid in have or pid not in sproj:
            continue
        nm = (players.get(pid) or {}).get("full_name") or pid
        board.append({"id": pid, "name": nm, "pos": pos, "team": r.get("team"),
                      "bye": None, "low": None, "proj": sproj[pid], "high": None,
                      "miss": None, "tier": None, "adp": adp.get(pid)})

    by_pos = {}
    for b in board:
        by_pos.setdefault(b["pos"], []).append(b)
    repl, used = replacement_levels(by_pos)
    for b in board:
        b["vbd"] = round(b["proj"] - repl.get(b["pos"], 0), 1)
        b["proj"] = round(b["proj"], 1)
        for k in ("low", "high"):
            if isinstance(b[k], (int, float)):
                b[k] = round(b[k], 1)
        if isinstance(b["miss"], (int, float)):
            b["miss"] = round(b["miss"], 2)

    board.sort(key=lambda x: -x["vbd"])
    out = {"generated": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
           "replacement": {k: round(v, 1) for k, v in repl.items()},
           "starters_filled": used, "players": board}
    OUT.write_text(json.dumps(out, separators=(",", ":")))
    print(f"wrote {OUT}  players={len(board)}")
    print("replacement level:", out["replacement"])
    print(f"unmatched from workbook ({len(unmatched)}):", ", ".join(unmatched[:15]) or "none")


if __name__ == "__main__":
    main()
